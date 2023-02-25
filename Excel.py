from typing import Iterable
from openpyxl import Workbook, load_workbook, styles
from string import ascii_uppercase


class Excel:
    """A class that helps you by its functions to work easier with Excel data

    Args:
        file (str): file name or file path + 'xlsx'
            Tip: if file not exist its will be create.
    """

    def __init__(self, file):
        try:
            self.file = file
            self.workbook = load_workbook(file)
        except FileNotFoundError:
            self.file = file
            self.workbook = Workbook()
            self.workbook.save(file)

    def _excel_columns(self, start='A'):
        """Internal generator

        Generate excel column: [A, B ... AA, AB ... XFE, XFD]
        """
        break_check = False
        start_check = False
        alphabet = list(ascii_uppercase)
        start_len = len(start)
        if start_len == 1:
            for item in alphabet:
                if start_check is False and start == item: start_check = True
                if start_check: yield item
        if start_len == 1 or start_len == 2:
            for item_1 in alphabet:
                for item_2 in alphabet:
                    if start_check is False and item_1 + item_2 == start: start_check = True
                    if start_check: yield item_1 + item_2
        for item_1 in alphabet:
            if break_check:
                break
            for item_2 in alphabet:
                if break_check:
                    break
                for item_3 in alphabet:
                    if (item_1 + item_2 + item_3) == "XFE":
                        break_check = True
                        break
                    if start_check is False and item_1 + item_2 + item_3 == start: start_check = True
                    if start_check: yield item_1 + item_2 + item_3

    def _excel_column_index(self, column):
        """Internal function

        Get index of columns of Excel
        """
        counter = 0
        for column_ in self._excel_columns():
            counter += 1
            if column_ == column:
                return counter

    def _sheet_config(self, sheet, read=False):
        """Internal function

        Config sheet of workbook
        if read = True it configs sheet for read functions
        """
        try:
            sheet = self.workbook[sheet]
        except KeyError:
            if read:
                raise KeyError(f"Worksheet {sheet} does not exist")
            sheets = self.workbook.sheetnames
            if "Sheet" in sheets:
                default = self.workbook["Sheet"]
                self.workbook.remove(default)
            self.workbook.create_sheet(sheet)
            sheet = self.workbook[sheet]
        return sheet

    def _column_verify(self, column):
        """Internal function

        Verify column name
        """
        if column.upper() not in self._excel_columns():
            raise KeyError("Column letter must be between A, B ... AA, AB and XFD")

    def _row_verify(self, row):
        """Internal function

        Verify row index
        """
        if row < 1 or row > 1048576:
            raise KeyError("Row index must be integer between 1 and 1048576")

    def write_on_column(
        self, sheet: str, column: str, values: Iterable, row_start: int = 1, center_style: bool = False
    ):
        """Write some values of a Iterable on a column

        Args:
            sheet (str): Sheet name. if it not exist will be create
            column (str): an excel column name. Ex AB **between A and XFD**
            values (iterable): Contains values that will fill cells
            row_start (int) [optional]: an Excel row index that starts the function from there. Ex 12 **between 1 and 1048576** (default = 1)
            center_style (bool) [optional]: if equal True styles of the cells will be middle (default = False)

        Return:
            True: if commands are properly executed
        """
        sheet = str(sheet)
        self._column_verify(column)
        sheet = self._sheet_config(sheet)

        for index, item in enumerate(values):
            sheet[f"{column.upper()}{index + row_start}"] = item
            if center_style:
                sheet[f"{column.upper()}{index + row_start}"].alignment = styles.Alignment(
                    horizontal="center", vertical="center"
                )
        self.workbook.save(self.file)
        return True

    def write_on_row(
        self, sheet: str, row: int, values: Iterable, col_start: str = 'A', center_style: bool = False
    ):
        """Write some values of a Iterable on a row

        Args:
            sheet (str): Sheet name. if it not exist will be create
            row (int): an excel row index. Ex 12 **between 1 and 1048576**
            values (iterable): Contains values that will fill cells
            col_start (str) [optional]: an Excel column that starts the function from there. Ex AB **between A and XFD** (default = A)
            center_style (bool) [optional]: if equal True styles of the cells will be middle (default = False)

        Return:
            True: if commands are properly executed
        """
        sheet = str(sheet)
        self._row_verify(row)
        sheet = self._sheet_config(sheet)

        for item, column in zip(values, self._excel_columns(start=col_start)):
            sheet[f"{column.upper()}{row}"] = item
            if center_style:
                sheet[f"{column.upper()}{row}"].alignment = styles.Alignment(
                    horizontal="center", vertical="center"
                )

        self.workbook.save(self.file)
        return True

    def write_on_cell(
        self, sheet: str, column: str, row: int, value, center_style: bool = False
    ):
        """Write a value on a cell

        Args:
            sheet (str): Sheet name. if it not exist will be create
            column (str): an excel column name. Ex AB **between A and XFD**
            row (int): an excel row index. Ex 12 **between 1 and 1048576**
            value (any except Iterables): value that will fill cells
            center_style (bool) [optional]: if equal True styles of the cells will be middle (default = False)

        Return:
            True: if commands are properly executed
        """
        sheet = str(sheet)
        self._column_verify(column)
        self._row_verify(row)
        sheet = self._sheet_config(sheet)

        sheet[f"{column}{row}"] = value
        if center_style:
            sheet[f"{column}{row}"].alignment = styles.Alignment(
                horizontal="center", vertical="center"
            )
        self.workbook.save(self.file)
        return True

    def read_column(self, sheet: str, column: str, row_start: int = 1, row_end: int = 1048576):
        """Read some values of an excel column

        Args:
            sheet (str): Sheet name
            column (str): an excel column name. Ex AB **between A and XFD**
            row_start (int) [optional]: an Excel row index that starts the generator from there. Ex 12 **between 1 and 1048576** (default = 1)
            row_end (int) [optional]: an Excel row index that breake the generator from there. Ex 12 **between 1 and 1048576** (default = 1048576)

        Yields:
            Column values
        """
        sheet = str(sheet)
        self._column_verify(column)
        sheet = self._sheet_config(sheet, True)

        generator = sheet.values
        col_index = self._excel_column_index(column)
        for index,row in enumerate(generator):
            if (index + 1) < row_start : continue
            if (index + 1) > row_end : break
            yield row[col_index - 1]

    def read_row(self, sheet: str, row: int, col_start: str = 'A', col_end: str = 'XFD'):
        """Read some values of an excel row

        Args:
            sheet (str): Sheet name
            row (int): an excel row index. Ex 12 **between 1 and 1048576**
            col_start (str) [optional]: an Excel column that starts the generator from there. Ex AB **between A and XFD** (default = A)
            col_end (str) [optional]: an Excel column that breake the generator from there. Ex AB **between A and XFD** (default = XFD)
        Yields:
            Row values
        """
        sheet = str(sheet)
        self._row_verify(row)
        sheet = self._sheet_config(sheet, True)

        generator = sheet.values
        col_index_start = self._excel_column_index(col_start)
        col_index_end = self._excel_column_index(col_end)
        counter = 0
        for row_ in generator:
            counter += 1
            if counter == row:
                for col_index,cell in enumerate(row_):
                    if (col_index + 1) < col_index_start : continue
                    if (col_index + 1) > col_index_end : break
                    yield cell

    def read_cell(self, sheet: str, column: str, row: int):
        """Read value of an excel cell

        Args:
            sheet (str): Sheet name
            column (str): an excel column name. Ex AB **between A and XFD**
            row (int): an excel row index. Ex 12 **between 1 and 1048576**

        Return:
            The cell value
        """
        sheet = str(sheet)
        self._column_verify(column)
        self._row_verify(row)
        sheet = self._sheet_config(sheet)

        generator = sheet.values
        column = self._excel_column_index(column) - 1
        counter = 0
        for row_ in generator:
            counter += 1
            if counter == row:
                return row_[column]
